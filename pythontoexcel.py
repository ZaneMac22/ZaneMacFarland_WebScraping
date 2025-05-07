import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title="Second Sheet")

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman',size = 24, bold=True)

fontobject = Font(name='Times New Roman',size = 24, bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'engine'


wb.save('PythontoExcel.xlsx')

